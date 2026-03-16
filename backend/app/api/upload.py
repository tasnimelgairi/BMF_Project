from fastapi import APIRouter, UploadFile, File, Form, HTTPException
from pathlib import Path
import shutil
from openpyxl import load_workbook
import pandas as pd
from datetime import datetime
from app.core.database import get_connection
router = APIRouter()

BASE_DIR = Path(__file__).resolve().parents[3]
INPUT_DIR = BASE_DIR / "data" / "input"
REFERENCE_FILE = BASE_DIR / "data" / "reference" / "reference_template.xlsm"

INPUT_DIR.mkdir(parents=True, exist_ok=True)
REFERENCE_FILE.parent.mkdir(parents=True, exist_ok=True)


def normalize_text(value: str) -> str:
    return " ".join(str(value).strip().lower().split())


@router.post("/upload")
async def upload_file(
    file: UploadFile = File(...),
    supplier: str = Form(...),
    run_date: str = Form(...)
):
    if not file.filename:
        raise HTTPException(status_code=400, detail="Aucun fichier sélectionné.")

    file_path = INPUT_DIR / file.filename

    try:
        if not REFERENCE_FILE.exists():
            raise HTTPException(
                status_code=500,
                detail="Le fichier de référence est introuvable."
            )

        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        workbook = load_workbook(file_path, read_only=True)
        sheet_names = workbook.sheetnames

        expected_sheet_1 = "Supplier all countries values"
        expected_sheet_2 = "Supplier values by country"

        has_sheet_1 = expected_sheet_1 in sheet_names
        has_sheet_2 = expected_sheet_2 in sheet_names

        if not has_sheet_1 or not has_sheet_2:
            raise HTTPException(
                status_code=400,
                detail={
                    "message": "Les feuilles attendues n'existent pas.",
                    "sheet_count": len(sheet_names),
                    "detected_sheets": sheet_names,
                    "expected_sheets": [expected_sheet_1, expected_sheet_2],
                    "supplier_all_countries_values_exists": has_sheet_1,
                    "supplier_values_by_country_exists": has_sheet_2
                }
            )

        validation_result = {}

        for sheet_name in [expected_sheet_1, expected_sheet_2]:
            input_df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=0)
            reference_df = pd.read_excel(REFERENCE_FILE, sheet_name=sheet_name, nrows=0)

            input_columns = input_df.columns.tolist()
            reference_columns = reference_df.columns.tolist()

            normalized_input_columns = [normalize_text(col) for col in input_columns]
            normalized_reference_columns = [normalize_text(col) for col in reference_columns]

            missing_columns = [
                reference_columns[i]
                for i, col in enumerate(normalized_reference_columns)
                if col not in normalized_input_columns
            ]

            extra_columns = [
                input_columns[i]
                for i, col in enumerate(normalized_input_columns)
                if col not in normalized_reference_columns
            ]

            same_columns = set(normalized_input_columns) == set(normalized_reference_columns)

            validation_result[sheet_name] = {
                "reference_columns": reference_columns,
                "input_columns": input_columns,
                "missing_columns": missing_columns,
                "extra_columns": extra_columns,
                "is_valid": same_columns
            }

            if not same_columns:
                raise HTTPException(
                    status_code=400,
                    detail={
                        "message": f"La feuille '{sheet_name}' ne contient pas les mêmes colonnes que le fichier de référence.",
                        "missing_columns": missing_columns,
                        "extra_columns": extra_columns,
                        "reference_columns": reference_columns,
                        "input_columns": input_columns
                    }
                )
            columns_by_sheet = {}

            for sheet_name in [expected_sheet_1, expected_sheet_2]:
                df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=0)
                columns_by_sheet[sheet_name] = df.columns.tolist()

        conn = get_connection()

        conn.execute("""
            INSERT INTO uploaded_files (
                file_name,
                file_path,
                supplier,
                run_date,
                status,
                uploaded_at
            )
            VALUES (?, ?, ?, ?, ?, ?)
        """, [
            file.filename,
            str(file_path),
            supplier,
            run_date,
            "validated",
            datetime.now()
        ])


        conn.close()   
        return {
            "message": "Les feuilles existent et les colonnes correspondent au fichier de référence.",
            "sheet_count": len(sheet_names),
            "sheet_names": sheet_names,
            "supplier_all_countries_values_exists": has_sheet_1,
            "supplier_values_by_country_exists": has_sheet_2,
            "validation": validation_result,
            "columns_by_sheet": columns_by_sheet
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Erreur lors de la lecture du fichier : {str(e)}"
        )